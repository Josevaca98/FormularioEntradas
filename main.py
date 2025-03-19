# Importamos las librerías necesarias
import tkinter as tk  # Para la interfaz gráfica
from tkinter import messagebox  # Para mostrar mensajes emergentes
from openpyxl import Workbook, load_workbook  # Para manejar archivos Excel
import re  # Para validación de correo electrónico
import os  # Para comprobar si el archivo ya existe

# Nombre del archivo de Excel donde se guardarán los datos
nombre_archivo = 'datos.xlsx'

# Comprobamos si el archivo ya existe
if os.path.exists(nombre_archivo):
    # Si existe, lo abrimos para seguir guardando datos
    wb = load_workbook(nombre_archivo)
    ws = wb.active
else:
    # Si no existe, creamos un nuevo archivo de Excel
    wb = Workbook()
    ws = wb.active
    # Agregamos la fila de encabezados
    ws.append(["Nombre", "Edad", "Email", "Teléfono", "Dirección"])

# Función para guardar los datos ingresados en el formulario
def guardar_datos():
    # Obtenemos los valores de los campos de entrada
    nombre = entry_nombre.get()
    edad = entry_edad.get()
    email = entry_email.get()
    telefono = entry_telefono.get()
    direccion = entry_direccion.get()

    # Verificamos que todos los campos estén llenos
    if not nombre or not edad or not email or not telefono or not direccion:
        messagebox.showwarning("Advertencia", "Todos los campos son obligatorios")
        return
    
    # Intentamos convertir la edad y el teléfono en números enteros
    try:
        edad = int(edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showwarning("Advertencia", "Edad y Teléfono deben ser números")
        return
    
    # Validamos el formato del email con una expresión regular
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showwarning("Advertencia", "El correo electrónico no es válido")
        return
    
    # Agregamos los datos a la hoja de Excel
    ws.append([nombre, edad, email, telefono, direccion])
    wb.save(nombre_archivo)  # Guardamos el archivo

    # Mostramos un mensaje de confirmación
    messagebox.showinfo("Información", "Datos guardados con éxito")

    # Limpiamos los campos de entrada para nuevos registros
    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)
    entry_direccion.delete(0, tk.END)

# Crear la ventana principal de la aplicación
root = tk.Tk()
root.title("Formulario de Entrada de Datos")  # Título de la ventana
root.configure(bg='#020611')  # Color de fondo

# Definimos los estilos para etiquetas y entradas de texto
label_style = {"bg": '#020611', "fg": "#f4feff"}  # Estilo de etiquetas
entry_style = {"bg": '#494cd1', "fg": "#f4feff"}  # Estilo de campos de entrada

# Campo Nombre
label_nombre = tk.Label(root, text="Nombre", **label_style)
label_nombre.grid(row=0, column=0, padx=10, pady=5)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)

# Campo Edad
label_edad = tk.Label(root, text="Edad", **label_style)
label_edad.grid(row=1, column=0, padx=10, pady=5)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=5)

# Campo Email
label_email = tk.Label(root, text="E-mail", **label_style)
label_email.grid(row=2, column=0, padx=10, pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=10, pady=5)

# Campo Teléfono
label_telefono = tk.Label(root, text="Teléfono", **label_style)
label_telefono.grid(row=3, column=0, padx=10, pady=5)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=10, pady=5)

# Campo Dirección
label_direccion = tk.Label(root, text="Dirección", **label_style)
label_direccion.grid(row=4, column=0, padx=10, pady=5)
entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=4, column=1, padx=10, pady=5)

# Botón para guardar los datos
boton_guardar = tk.Button(root, text="Guardar", command=guardar_datos, bg='#343b57', fg="white")
boton_guardar.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

# Ejecutamos la aplicación
root.mainloop()
